from datetime import datetime

from PyQt5 import uic, QtGui
from PyQt5 import QtWidgets, QtSql, QtCore
from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


# создается класс формы Класса ЧС
class ReportForm(QtWidgets.QDialog):
    def __init__(self):
        super(ReportForm, self).__init__()
        uic.loadUi("./ui/Report.ui", self)
        self.setWindowModality(2)
        self.ui_init()

    def ui_init(self):
        #self.dateEdit1.dateChanged.connect(self.DateEChanged)
        #self.dateEdit2.dateChanged.connect(self.DateEChanged)
        self.tWReport.horizontalHeader().setDefaultSectionSize(150)
        self.tWReport.setColumnCount(4)
        self.tWReport.setHorizontalHeaderLabels(["Виды ЧС","Количество ЧС","Пострадало","Из них погибло"])
        self.tWReport.resizeColumnsToContents()
        self.ComboBox.addItem("Виды ЧС")
        self.ComboBox.addItem("Количество ЧС")
        self.ComboBox.addItem("Пострадало")
        self.ComboBox.addItem("Из них погибло")
        self.ComboBox.currentIndexChanged.connect(self.ComBox_changed)
        self.CheckB.stateChanged.connect(self.Desc_changed)
        self.pBtnReport.clicked.connect(self.btnClickR)
        self.pBtnClose.clicked.connect(self.btnClickClose)

        self.spinBox.valueChanged.connect(self.DEChanged)
        self.spinBox.setValue(2021)



    def report_get_results(self):
        self.report_query = QtSql.QSqlQuery()
        self.report_query.prepare(
          "SELECT vid.name2, "
          "COUNT(chs.id_Vid) AS CountVid, "
          "SUM(chs.postradalo) AS SummaPostradav, "
          "SUM(chs.pogiblo) AS SummaPogiblo "
          "FROM (vid JOIN chs ON (vid.id_Vid=chs.id_Vid)) "
          #"WHERE (chs.Data >= :dateEdit1) AND (chs.Data <= :dateEdit2) "
          f"WHERE (chs.Data >= '{self.spinBox.value()}.01.01') AND (chs.Data <= '{self.spinBox.value()}.12.31') "
          "GROUP BY vid.name2 "
          f"ORDER BY {'vid.name2' if self.ComboBox.currentIndex()==0 else 'CountVid' if self.ComboBox.currentIndex()==1 else 'SummaPostradav' if self.ComboBox.currentIndex()==2 else 'SummaPogiblo'  } "
          f"{'DESC' if self.CheckB.isChecked() else 'ASC'}"
        )

        self.report_query.exec()
        self.tWReport.setRowCount(self.report_query.size())
        i = 0
        while self.report_query.next():
            for j in range(self.tWReport.columnCount()):
                item = QtWidgets.QTableWidgetItem(str(self.report_query.value(j)))
                if j==0:
                    self.tWReport.setItem(i, j, item)
                if j == 1:
                    #if self.report_query.value(j) >= 1:
                     #   item.setBackground(QtGui.QColor(255,128,128,255))
                    self.tWReport.setItem(i, j, item)
                if j == 2:
                    #if self.report_query.value(j) >= 1:
                     #   item.setBackground(QtGui.QColor(155,122,150,111))
                    self.tWReport.setItem(i, j, item)
                if j == 3:
                    #if self.report_query.value(j) >= 1:
                     #   item.setBackground(QtGui.QColor(234, 111, 176, 100))
                    self.tWReport.setItem(i, j, item)

            i += 1


    def DateEChanged(self):
        self.report_get_results()

    def DEChanged(self):
        self.report_get_results()

    def ComBox_changed(self):
        self.report_get_results()

    def Desc_changed(self):
        self.report_get_results()

    def btnClickClose(self):
        self.close()

    def btnClickR(self):
        wb = load_workbook('./templates/Сводные данные по ЧС.xlsx')
        ws = wb["Sheet1"]

        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        )
        ws["E2"].value = f"c 01.01 " \
                         f" по 12.31.{self.spinBox.value()}года"
        ws["C5"].value = f"{self.spinBox.value()}"
        ws["E5"].value = f"{self.spinBox.value()}"
        ws["G5"].value = f"{self.spinBox.value()}"
        ws["D5"].value = f"{self.spinBox.value()-1}"
        ws["F5"].value = f"{self.spinBox.value()-1}"
        ws["H5"].value = f"{self.spinBox.value()-1}"
        for i in range(self.tWReport.rowCount()):
            ws.cell(i + 6, 1).value = i + 1
            ws.cell(i + 6, 1).border = border
            for j in range(self.tWReport.columnCount()):
                if j == 1:
                  ws.cell(i + 6, 3).value = float(self.tWReport.item(i, j).text())
                elif j == 2:
                  ws.cell(i + 6, 5).value = float(self.tWReport.item(i, j).text())
                elif j == 3:
                  ws.cell(i + 6, 7).value = float(self.tWReport.item(i, j).text())
                else:
                  ws.cell(i + 6, j + 2).value = self.tWReport.item(i, j).text()
                ws.cell(i + 6, j + 2).border = border

        self.report_query = QtSql.QSqlQuery()
        self.report_query.prepare(
            "SELECT vid.name2, "
            "COUNT(chs.id_Vid) AS CountVid, "
            "SUM(chs.postradalo) AS SummaPostradav, "
            "SUM(chs.pogiblo) AS SummaPogiblo "
            "FROM (vid JOIN chs ON (vid.id_Vid=chs.id_Vid)) "
            f"WHERE (chs.Data >= '{self.spinBox.value()-1}.01.01') AND (chs.Data <= '{self.spinBox.value()-1}.12.31') "
            "GROUP BY vid.name2 "
            f"ORDER BY {'vid.name2' if self.ComboBox.currentIndex() == 0 else 'CountVid' if self.ComboBox.currentIndex() == 1 else 'SummaPostradav' if self.ComboBox.currentIndex() == 2 else 'SummaPogiblo'} "
            f"{'DESC' if self.CheckB.isChecked() else 'ASC'}"
        )
        self.report_query.exec()

        i = 0
        while self.report_query.next():
            for j in range(4):
                item = str(self.report_query.value(j))
                if j == 1:
                  ws.cell(i + 6, 4).value = float(item)
                  ws.cell(i + 6, 4).border = border
                elif j == 2:
                  ws.cell(i + 6, 6).value = float(item)
                  ws.cell(i + 6, 6).border = border
                elif j == 3:
                  ws.cell(i + 6, 8).value = float(item)
                  ws.cell(i + 6, 8).border = border
            i += 1
        filename = QFileDialog.getSaveFileName(
            self,
            "Сохранение отчета",
            "Сводные данные о ЧС_",
            "MS Excel Files(*.xlsx *xls)"
        )

        if filename[0]:
            wb.save(filename[0])